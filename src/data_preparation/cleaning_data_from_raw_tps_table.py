""" This script is used to clean the raw TPS excel table """
import pickle

import pandas as pd  # type: ignore
import numpy as np  # type: ignore
from openpyxl import load_workbook  # type: ignore
from rdkit import Chem  # type: ignore
from src.utils.data import get_canonical_smiles

# A dictionary mapping the excel substrates to SMILES strings.
SUBSTRATE_NAME_2_SMILES = {
    "(2E,6E)-FPP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O",
    "(2E,6E,10E)-GGPP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O",
    "GGPP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O",
    "(+)-copalyl diphosphate": "[H][C@@]12CCC(=C)[C@H](CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O)[C@@]1(C)CCCC2(C)C",
    "(2E,6E)-FPP; isopentenyl PP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O.CC(=C)CCOP(=O)(O)OP(=O)(O)O",
    "(2E)-GPP; isopentenyl PP": "CC(C)=CCC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O.CC(=C)CCOP(=O)(O)OP(=O)(O)O",
    "dimethylallyl PP; isopentenyl PP": "CC(=CCOP(=O)([O-])OP(=O)([O-])[O-])C.CC(=C)CCOP(=O)(O)OP(=O)(O)O",
    "(2E)-GPP": "CC(C)=CCC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O",
    "ent-copalyl diphosphate": "[C@@H]1(CC/C(/C)=C/COP(OP(=O)([O-])[O-])(=O)[O-])C(=C)CC[C@]2([C@@]1(C)CCCC2(C)C)[H]",
    "(2Z,6E)-FPP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C/COP([O-])(=O)OP([O-])([O-])=O",
    "(S)-2,3-epoxysqualene": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\CC\\C=C(/C)CC\\C=C(/C)CC[C@@H]1OC1(C)C",
    "peregrinol PP": "[C@@]1(CC/C(/C)=C/COP(OP(=O)([O-])[O-])(=O)[O-])([C@@H](CC[C@@]2([C@]1(C)CCCC2(C)C)[H])C)O",
    "copal-8-ol diphosphate(3−)": "[H][C@@]12CC[C@@](C)(O)[C@H](CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O)[C@@]1(C)CCCC2(C)C",
    "NPP": "CC(=CCC/C(=C\\COP(=O)(O)OP(=O)(O)O)/C)C",
    "9α-copalyl PP": "[H][C@@]12CCC(=C)[C@@H](CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O)[C@@]1(C)CCCC2(C)C",
    "(2Z,6Z)-FPP": "CC(C)=CCC\\C(C)=C/CC\\C(C)=C/COP([O-])(=O)OP([O-])([O-])=O",
    "ent-copal-8-ol diphosphate(3−)": "[C@@H]1(CC/C(/C)=C/COP(OP(=O)([O-])[O-])(=O)[O-])[C@](CC[C@]2([C@@]1(C)CCCC2(C)C)[H])(O)C",
    "9α-copalyl diphosphate": "[H][C@@]12CCC(=C)[C@@H](CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O)[C@@]1(C)CCCC2(C)C",
    "(2E,6E,10E,14E)-GFPP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O",
    "(R)-tetraprenyl-β-curcumene": "C[C@H](CC\\C=C(/C)CC\\C=C(/C)CC\\C=C(/C)CC\\C=C(/C)CCC=C(C)C)C1=CCC(C)=CC1",
    "(E)-2-MeGPP": "CC(C)=CCC\\C(C)=C(/C)COP([O-])(=O)OP([O-])([O-])=O",
    "dimethylallyl PP; 4 isopentenyl PP": "CC(=C)CCOP(=O)(O)OP(=O)(O)O.CC(=C)CCOP(=O)(O)OP(=O)(O)O.CC(=C)CCOP(=O)(O)OP(=O)(O)O.CC(=C)CCOP(=O)(O)OP(=O)(O)O.CC(=CCOP(=O)([O-])OP(=O)([O-])[O-])C",
    "(3S,22S)-2,3:22,23-diepoxy-2,3,22,23-tetrahydrosqualene": "C(C/C=C(/CC/C=C(/CC[C@H]1C(O1)(C)C)\\C)\\C)/C=C(/CC/C=C(/CC[C@H]2C(O2)(C)C)\\C)\\C",
    "pre-α-onocerin": "C1C[C@@H](C([C@]2([C@]1([C@H](C(CC2)=C)CC/C=C(/CC/C=C(/CC[C@H]3C(O3)(C)C)\\C)\\C)C)[H])(C)C)O",
    "dimethylallyl PP": "CC(=CCOP(=O)([O-])OP(=O)([O-])[O-])C",
    "(R,R)-chrysanthemyl diphosphate": "CC(C)=C[C@@H]1[C@@H](COP([O-])(=O)OP([O-])([O-])=O)C1(C)C",
    "(R)-lavandulyl diphosphate(3−)": "C(=CC[C@@H](COP([O-])(=O)OP(=O)([O-])[O-])C(C)=C)(C)C",
    "PPP": "CC(C)CCC[C@@H](C)CCC[C@@H](C)CCC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O",
    "HexPP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O",
    "HepPP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O",
    "presqualene PP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\[C@H]1[C@H](COP([O-])(=O)OP([O-])([O-])=O)[C@@]1(C)CC\\C=C(/C)CCC=C(C)C",
    "isopentenyl PP; (2E,6E,10E)-GGPP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O.CC(=C)CCOP(=O)(O)OP(=O)(O)O",
    "isopentenyl PP; (2E,6E)-FPP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O.CC(=C)CCOP(=O)(O)OP(=O)(O)O",
    "isopentenyl PP; (2E)-GPP": "CC(C)=CCC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O.CC(=C)CCOP(=O)(O)OP(=O)(O)O",
    "prephytoene PP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\C1C(COP(O)(=O)OP(O)(O)=O)C1(C)CC\\C=C(/C)CC\\C=C(/C)CCC=C(C)C",
    "(2E,6E,10E)-GGPP; isopentenyl PP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O.CC(=C)CCOP(=O)(O)OP(=O)(O)O",
    "Unknown": "Unknown",
    "isopentenyl PP": "CC(=C)CCOP(=O)(O)OP(=O)(O)O",
    "γ-carotene": "CC1=C(C(CCC1)(C)C)C=CC(=CC=CC(=CC=CC=C(C)C=CC=C(C)C=CC=C(C)CCC=C(C)C)C)C",
    "CDP": "CC(=CCOP(=O)(O)OP(=O)(O)O)CCC1C(=C)CCC2C1(CCCC2(C)C)C",
    "(2E,6E)-FPP + IPP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O.CC(=C)CCOP([O-])(=O)OP([O-])([O-])=O",
    "(2E)-GPP + IPP": "CC(C)=CCC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O.CC(=C)CCOP([O-])(=O)OP([O-])([O-])=O",
    "dimethylallyl PP+ IPP": "CC(=CCOP(=O)([O-])OP(=O)([O-])[O-])C.CC(=C)CCOP(=O)(O)OP(=O)(O)O",
    "squalene": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\CC\\C=C(/C)CC\\C=C(/C)CCC=C(C)C",
    "(13E)-labda-7,13-dien-15-yl PP": "[C@H]1(CC/C(/C)=C/COP(OP(=O)(O)O)(=O)O)C(C)=CC[C@@]2([C@]1(C)CCCC2(C)C)[H]",
    "all-trans-lycopene": "CC(C)=CCC\\C(C)=C\\C=C\\C(C)=C\\C=C\\C(C)=C\\C=C\\C=C(C)\\C=C\\C=C(C)\\C=C\\C=C(/C)CCC=C(C)C",
    "dimethylallyl PP + isopentenyl PP": "CC(=CCOP(=O)([O-])OP(=O)([O-])[O-])C.CC(=C)CCOP(=O)(O)OP(=O)(O)O",
    "(2E,6E)-FPP;  isopentenyl PP": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O.CC(=C)CCOP([O-])(=O)OP([O-])([O-])=O",
    "(2E,6E)-FPP ": "CC(C)=CCC\\C(C)=C\\CC\\C(C)=C\\COP([O-])(=O)OP([O-])([O-])=O",
    "syn-CPP": "C/C(=C\\COP(=O)(O)OP(=O)(O)O)/CC[C@@H]1C(=C)CC[C@H]2[C@]1(CCCC2(C)C)C",
    "CPP": "C/C(=C\\COP(=O)(O)OP(=O)(O)O)/CC[C@@H]1C(=C)CC[C@H]2[C@]1(CCCC2(C)C)C",
    "syn-CDP": "CC(=CCOP(=O)(O)OP(=O)(O)O)CCC1C(=C)CCC2C1(CCCC2(C)C)C",
    "(2Z,6Z,10Z)-NNPP": "CC(C)=CCCC(/C)=C/CCC(/C)=C/CCC(/C)=C/COP([O-])(OP([O-])([O-])=O)=O",
}


if __name__ == "__main__":
    # read the main excel table
    df_main = pd.read_excel("data/TPS-database_Nov19_23.xlsx")

    # cleaning the database based on manual comments about the entries
    workbook = load_workbook("data/TPS-database_Nov19_23.xlsx")
    sheet1 = workbook.sheetnames[0]
    worksheet = workbook[sheet1]

    blacklisted_not_characterized = set()
    for row in worksheet.iter_rows():
        cell = row[0]
        if cell.comment:
            for problematic_str in [
                "not characterized",
                "not characetrized",
                "could not verify",
                "not find",
                "ques-tion mark also remains over the direct products",
                "wrongly annotated",
                "full length protein was not active",
            ]:
                if problematic_str in cell.comment.text:
                    blacklisted_not_characterized.add(cell.value)
                    break

    df_main = df_main[~df_main["Uniprot ID"].isin(blacklisted_not_characterized)]
    df_main.loc[
        df_main["Type (mono, sesq, di, …)"] == "Negative",
        "Substrate (including stereochemistry)",
    ] = "Negative"
    SUBSTRATE_NAME_2_SMILES.update({"Negative": "Negative"})

    df_main["OK row flag"] = (
        ~df_main[
            [
                "Amino acid sequence",
                "Substrate (including stereochemistry)",
            ]
        ]
        .isnull()
        .any(axis=1)
        .astype(int)
    )
    df_main.loc[~df_main["Fragment"].isnull(), "OK row flag"] = 0

    df_main["SMILES of substrate"] = df_main[
        "Substrate (including stereochemistry)"
    ].map(SUBSTRATE_NAME_2_SMILES)
    df_main["SMILES of substrate"] = df_main["SMILES of substrate"].fillna("Unknown")

    df_main["SMILES of substrate"] = df_main[
        "Substrate (including stereochemistry)"
    ].map(SUBSTRATE_NAME_2_SMILES)
    df_main = df_main[
        np.logical_not(df_main["SMILES of substrate"].isnull())
        | (df_main["Type (mono, sesq, di, …)"] == "Negative")
    ]

    # adding hard negatives (single-point mutants without TPS activity)
    col_2_vals = {
        "Uniprot ID": ["Q9PDX7", "P0C0L4", "P06238", "P20742", "Q9GYW4"],
        "Amino acid sequence": [
            "".join(
                """MRDRVAMMLRPLVRGWIPRAVLLLTVAFSFGCNRNHNGQLPQSSGEPVAVAKEPVKGFVL
    VRAYPDQHDGELALALEFSQPLAATQEFDTLVRLEQDSGNHDGGWSLSDDAKTLRYPYVE
    ADKHYTVLISAGLLAATGSRLGKPRKEPVYTGELDPVVGFASRGSILPARGSRGVPVVSV
    NVPEVDVEFMRVREKALPAFLARYHKAGQRSSWELSNQGNSRKRLSELADPVYVTRFVLD
    GKKNERALTYLPIQSIRELREPGLYFAVMKPTGSFSDAFETAFFSVSNIGLHTRAYKDKL
    FVHTASLRSGNPYKQVDLLVLDAKGETVLQGATDDNGNALLNYTLNAGHVLVSRNGRDIS
    ILPFNQPALDLSEFAVAGRENPWFDVFAWSGRDLYRPGETLRISALLRDRDGKPVKPQPV
    FLRLKQPDGKTFRETRLQPAEQGYLEFTQKIPSDAPTGRWRVEFRTDPASKEAVQGLAVR
    VEEFLPERMKLELSSAQPVLRAKAPFTLTADAAYLYGAPAAGNRFTANLAVAVEQHPLDN
    MPGWFFGDATLQLPRGAKETIDITLGADGHLVHDIVLPEEAKPVSPMAVVVSGSVYESGG
    RPVTRSLKRVLWPADALVGVRPLFDVASGADANGMARFELTRVGVDGKPQSAKGLKATLV
    RELRDYHWRYSDGRWDYDFTRRFENKETRTVDISTSHTTTLALPVEWGDYRLEVFDPVTG
    LTMRYPFRAGWSWGDDNRGLDARPDKVKLALDKTSYRAGDTLKVTITPPHPGKGLLLVES
    DKPLYVQAIDANPSTTLEIPVTADWERHDVYVTALVFRGGSASNNTTPARAVGEVYVPMQ
    RKDRRVAVGLVVPKQMRPAQSLPVTVSVPELAGKQAHVTISAVDAGILNITGFPVPDAAA
    HFFAQRRLSVDAYDIYGRVIESFEGGTGRLKFGGDMALPPLPQAKRPTARSQTVDLFSGA
    VKLDAKGNAHIQLPVPDFNGALRVSALVYSDTRYGQRDAETVVRAPILAEASMPRVMAPG
    DRSTVTVDVQNFTGKQGKFAVKVEGVGPLAVAEAGRSVTLGIDGKTTLNFPLRALEGNSV
    AQVRVRVEGNGSKAERHYDLPVRAAWPQGLRTQAHVLNVLAPIAFDPALAKGLMPDSVNA
    RLSVSALAPIPFASVLQGVFEYPYGCAEQTASKGYAALWLDDATIKSLGIHGVTPAQRLE
    RLEGALGRLASLQTMNGHFSMCGGNSDVNPVLTPYIAGFLLDAKDAGFAVSDAVLQKALN
    RLSEDLLSGAHLFYGNDQSEALMFAHQAWSGYVLARVNRAPLGTLRTLYDNERGKAVSGL
    SLVHLGVALSLQGDRKRGEAAIEAGFAKSEGGRPEVFGDYGSVIRDNALMIALVRAHGLA
    KPAYEARVMALGRDLEARRRSGWLWLSTQEQVALAQLGRALLVDQKKQVSGTLYVGKQRE
    EIAASRLIGRSFDAAALARGVRFVPQGDVPLYASFEVAGIPRQAPVSDDSQLLVVRRWYT
    VDGKPWTPGPLKEGQALIVRVSVTSKQNMPDALLTDLLPAGLEIDNFNLGETRQWADVTV
    DGIALSERANAADIKHEEFRDDRYVAMLQLTGGRTANLFYLVRAVTPGTYKVPPSLVEDM
    YRPALRGTGRVAPATVTVVQP""".split()
            ),
            "".join(
                """MRLLWGLIWASSFFTLSLQKPRLLLFSPSVVHLGVPLSVGVQLQDVPRGQVVKGSVFLRN
    PSRNNVPCSPKVDFTLSSERDFALLSLQVPLKDAKSCGLHQLLRGPEVQLVAHSPWLKDS
    LSRTTNIQGINLLFSSRRGHLFLQTDQPIYNPGQRVRYRVFALDQKMRPSTDTITVMVEN
    SHGLRVRKKEVYMPSSIFQDDFVIPDISEPGTWKISARFSDGLESNSSTQFEVKKYVLPN
    FEVKITPGKPYILTVPGHLDEMQLDIQARYIYGKPVQGVAYVRFGLLDEDGKKTFFRGLE
    SQTKLVNGQSHISLSKAEFQDALEKLNMGITDLQGLRLYVAAAIIESPGGEMEEAELTSW
    YFVSSPFSLDLSKTKRHLVPGAPFLLQALVREMSGSPASGIPVKVSATVSSPGSVPEVQD
    IQQNTDGSGQVSIPIIIPQTISELQLSVSAGSPHPAIARLTVAAPPSGGPGFLSIERPDS
    RPPRVGDTLNLNLRAVGSGATFSHYYYMILSRGQIVFMNREPKRTLTSVSVFVDHHLAPS
    FYFVAFYYHGDHPVANSLRVDVQAGACEGKLELSVDGAKQYRNGESVKLHLETDSLALVA
    LGALDTALYAAGSKSHKPLNMGKVFEAMNSYDLGCGPGGGDSALQVFQAAGLAFSDGDQW
    TLSRKRLSCPKEKTTRKKRNVNFQKAINEKLGQYASPTAKRCCQDGVTRLPMMRSCEQRA
    ARVQQPDCREPFLSCCQFAESLRKKSRDKGQAGLQRALEILQEEDLIDEDDIPVRSFFPE
    NWLWRVETVDRFQILTLWLPDSLTTWEIHGLSLSKTKGLCVATPVQLRVFREFHLHLRLP
    MSVRRFEQLELRPVLYNYLDKNLTVSVHVSPVEGLCLAGGGGLAQQVLVPAGSARPVAFS
    VVPTAAAAVSLKVVARGSFEFPVGDAVSKVLQIEKEGAIHREELVYELNPLDHRGRTLEI
    PGNSDPNMIPDGDFNSYVRVTASDPLDTLGSEGALSPGGVASLLRLPRGCGEQTMIYLAP
    TLAASRYLDKTEQWSTLPPETKDHAVDLIQKGYMRIQQFRKADGSYAAWLSRDSSTWLTA
    FVLKVLSLAQEQVGGSPEKLQETSNWLLSQQQADGSFQDPCPVLDRSMQGGLVGNDETVA
    LTAFVTIALHHGLAVFQDEGAEPLKQRVEASISKANSFLGEKASAGLLGAHAAAITAYAL
    TLTKAPVDLLGVAHNNLMAMAQETGDNLYWGSVTGSQSNAVSPTPAPRNPSDPMPQAPAL
    WIETTAYALLHLLLHEGKAEMADQASAWLTRQGSFQGGFRSTQDTVIALDALSAYWIASH
    TTEERGLNVTLSSTGRNGFKSHALQLNNRQIRGLEEELQFSLGSKINVKVGGNSKGTLKV
    LRTYNVLDMKNTTCQDLQIEVTVKGHVEYTMEANEDYEDYEYDELPAKDDPDAPLQPVTP
    LQLFEGRRNRRRREAPKVVEEQESRVHYTVCIWRNGKVGLSGMAIADVTLLSGFHALRAD
    LEKLTSLSDRYVSHFETEGPHVLLYFDSVPTSRECVGFEAVQEVPVGLVQPASATLYDYY
    NPERRCSVFYGAPSKSRLLATLCSAEVCQCAEGKCPRQRRALERGLQDEDGYRMKFACYY
    PRVEYGFQVKVLREDSRAAFRLFETKITQVLHFTKDVKAAANQMRNFLVRASCRLRLEPG
    KEYLIMGLDGATYDLEGHPQYLLDSNSWIEEMPSERLCRSTRQRAACAQLNDFLQEYGTQ
    GCQV""".split()
            ),
            "".join(
                """MGKHRLRSLALLPLLLRLLLLLLPTDASAPQKPIYMVMVPSLLHAGTPEKACFLFSHLNE
    TVAVRVSLESVRGNQSLFTDLVVDKDLFHCTSFTVPQSSSDEVMFFTVQVKGATHEFRRR
    STVLVKKKESLVFAQTDKPIYKPGQTVRFRVVSLDESFHPLNELIPLLYIQDPKNNRIAQ
    WQNFNLEGGLKQLSFPLSSEPTQGSYKVVIRTESGRTVEHPFSVEEFVLPKFEVRVTVPE
    TITILEEEMNVSVCGIYTYGKPVPGRVTVNICRKYSNPSNCFGEESVAFCEKLSQQLDGR
    GCFSQLVKTKSFQLKRQEYEMQLDVHAKIQEEGTGVEETGKGLTKITRTITKLSFVNVDS
    HFRQGIPFVGQVLLVDGRGTPIPYETIFIGADEANLYINTTTDKHGLARFSINTDDIMGT
    SLTVRAKYKDSNACYGFRWLTEENVEAWHTAYAVFSPSRSFLHLESLPDKLRCDQTLEVQ
    AHYILNGEAMQELKELVFYYLMMAKGGIVRAGTHVLPLKQGQMRGHFSILISMETDLAPV
    ARLVLYAILPNGEVVGDTAKYEIENCLANKVDLVFRPNSGLPATRALLSVMASPQSLCGL
    RAVDQSVLLMKPETELSASLIYDLLPVKDLTGFPQGADQREEDTNGCVKQNDTYINGILY
    SPVQNTNEEDMYGFLKDMGLKVFTNSNIRKPKVCERLRDNKGIPAAYHLVSQSHMDAFLE
    SSESPTETRRSYFPETWIWDLVVVDSAGVAEVEVTVPDTITEWKAGAFCLSNDTGLGLSP
    VVQFQAFQPFFVELTMPYSVIRGEAFTLKATVLNYLPTCIRVAVQLEASPDFLAAPEEKE
    QRSHCICMNQRHTASWAVIPKSLGNVNFTVSAEALNSKELCGNEVPVVPEQGKKDTIIKS
    LLVEPEGLENEVTFNSLLCPMGAEVSELIALKLPSDVVEESARASVTVLGDILGSAMQNT
    QDLLKMPYGCGEQNMVLFAPNIYVLDYLNETQQLTQEIKTKAIAYLNTGYQRQLNYKHRD
    GSYSTFGDKPGRNHANTWLTAFVLKSFAQARKYIFIDEVHITQALLWLSQQQKDNGCFRS
    SGSLLNNAMKGGVEDEVTLSAYITIALLEMSLPVTHPVVRNALFCLDTAWKSARGGAGGS
    HVYTKALLAYAFALAGNQDTKKEILKSLDEEAVKEEDSVHWTRPQKPSVSVALWYQPQAP
    SAEVEMTAYVLLAYLTTEPAPTQEDLTAAMLIVKWLTKQQNSHGGFSSTQDTVVALHALS
    KYGSATFTRAKKAAQVTIHSSGTFSTKFQVNNNNQLLLQRVTLPTVPGDYTVKVTGEGCV
    YLQTSLKYSVLPREEEFPFTVVVQTLPGTCEDPKAHTSFQISLNISYTGSRSESNMAIAD
    VKMVSGFIPLKPTVKMLERSVHVSRTEVSNNHVLIYLDKVSNQTVNLSFTVQQDIPIRDL
    KPAVVKVYDYYEKDEFAVAKYSAPCSTDYGNA""".split()
            ),
            "".join(
                """MRKDRLLHLCLVLLLILLSASDSNSTEPQYMVLVPSLLHTEAPKKGCVLLSHLNETVTVS
    ASLESGRENRSLFTDLVAEKDLFHCVSFTLPRISASSEVAFLSIQIKGPTQDFRKRNTVL
    VLNTQSLVFVQTDKPMYKPGQTVRFRVVSVDENFRPRNELIPLIYLENPRRNRIAQWQSL
    KLEAGINQLSFPLSSEPIQGSYRVVVQTESGGRIQHPFTVEEFVLPKFEVKVQVPKIISI
    MDEKVNITVCGEYTYGKPVPGLATVSLCRKLSRVLNCDKQEVCEEFSQQLNSNGCITQQV
    HTKMLQITNTGFEMKLRVEARIREEGTDLEVTANRISEITNIVSKLKFVKVDSHFRQGIP
    FFAQVLLVDGKGVPIPNKLFFISVNDANYYSNATTNEQGLAQFSINTTSISVNKLFVRVF
    TVHPNLCFHYSWVAEDHQGAQHTANRVFSLSGSYIHLEPVAGTLPCGHTETITAHYTLNR
    QAMGELSELSFHYLIMAKGVIVRSGTHTLPVESGDMKGSFALSFPVESDVAPIARMFIFA
    ILPDGEVVGDSEKFEIENCLANKVDLSFSPAQSPPASHAHLQVAAAPQSLCALRAVDQSV
    LLMKPEAELSVSSVYNLLTVKDLTNFPDNVDQQEEEQGHCPRPFFIHNGAIYVPLSSNEA
    DIYSFLKGMGLKVFTNSKIRKPKSCSVIPSVSAGAVGQGYYGAGLGVVERPYVPQLGTYN
    VIPLNNEQSSGPVPETVRSYFPETWIWELVAVNSSGVAEVGVTVPDTITEWKAGAFCLSE
    DAGLGISSTASLRAFQPFFVELTMPYSVIRGEVFTLKATVLNYLPKCIRVSVQLKASPAF
    LASQNTKGEESYCICGNERQTLSWTVTPKTLGNVNFSVSAEAMQSLELCGNEVVEVPEIK
    RKDTVIKTLLVEAEGIEQEKTFSSMTCASGANVSEQLSLKLPSNVVKESARASFSVLGDI
    LGSAMQNIQNLLQMPYGCGEQNMVLFAPNIYVLNYLNETQQLTQEIKAKAVGYLITGYQR
    QLNYKHQDGSYSTFGERYGRNQGNTWLTAFVLKTFAQARSYIFIDEAHITQSLTWLSQMQ
    KDNGCFRSSGSLLNNAIKGGVEDEATLSAYVTIALLEIPLPVTNPIVRNALFCLESAWNV
    AKEGTHGSHVYTKALLAYAFSLLGKQNQNREILNSLDKEAVKEDNLVHWERPQRPKAPVG
    HLYQTQAPSAEVEMTSYVLLAYLTAQPAPTSGDLTSATNIVKWIMKQQNAQGGFSSTQDT
    VVALHALSRYGAATFTRTEKTAQVTVQDSQTFSTNFQVDNNNLLLLQQISLPELPGEYVI
    TVTGERCVYLQTSMKYNILPEKEDSPFALKVQTVPQTCDGHKAHTSFQISLTISYTGNRP
    ASNMVIVDVKMVSGFIPLKPTVKMLERSSSVSRTEVSNNHVLIYVEQVTNQTLSFSFMVL
    QDIPVGDLKPAIVKVYDYYETDESVVAEYIAPCSTDTEHGNV""".split()
            ),
            "".join(
                """MWQFIRSRILTVIIFIGAAHGLLVVGPKFIRANQEYTLVISNFNSQLSKVDLLLKLEGET
    DNGLSVLNVTKMVDVRRNMNRMINFNMPEELTAGNYKITIDGQRGFSFHKEAELVYLSKS
    ISGLIQVDKPVFKPGDTVNFRVILLDTELKPPARVKSVYVTIRDPQRNVIRKWSTAKLYA
    GVFESDLQIVPTPMLGVWNISVEVEGEELVSKTFEVKEYVLSTFDVQVMPSVIPLEEHQA
    VNLTIEANYHFGKPVQGVAKVELYLDDDKLNQKKELTVYGKGQVELRFDNFAMDADQQDV
    RVKVSFIEQYTNRTVVKQSQITVYRYAYRVELIKESPQFRPGLPFKCALQFTHHDGTPAK
    GITGKVEVSDVGFETTTTSDNDGLIKLELQPSEGTEQLGINFNAVDGFFFYEDVNKVETV
    TDAYIKLELKSPIKRNKLMRFMVTCTERMTFFVYYVMSKGNIIDAGFMRPNKQTKYLLQL
    NATEKMIPKAKILIATVAGRTVVYDYADLDFQELRNNFDLSIDEQEIKPGRQIELSMSGR
    PGAYVGLAAYDKALLLFNKNHDLFWEDIGQVFDGFHAINENEFDIFHSLGLFARTLDDIL
    FDSANEKTGRNALQSGKPIGKLVSYRTNFQESWLWKNVSIGRSGSRKLIEVVPDTTTSWY
    LTGFSIDPVYGLGIIKKPIQFTTVQPFYIVENLPYSIKRGEAVVLQFTLFNNLGAEYIAD
    VTLYNVANQTEFVGRPDTDLSYTKSVSVPPKVGVPISFLIKARKLGEMAVRVKASIMLGH
    ETDALEKVIRVMPESLAQPKMDTSFFCFDDYKNQTFPFNLDINKKADNGSKKIEFRLNPN
    LLTMVIKNLDNLLAVPTGCGEQNMVKFVPNILVLDYLYATGSKEQHLIDKATNLLRQGYQ
    NQMRYRQTDGSFGVWEKSGSSVFLTAFVATSMQTASKYMNDIDAAMVEKALDWLASKQHS
    SGRFDETGKVWHKDMQGGLRNGVALTSYVLTALLENDIAKVKHAVVIQNGMNYLSNQLAF
    INNPYDLSIATYAMMLNGHTMKKEALDKLIDMSISDNNKKERYWGTTNQIETTAYALLSF
    VMAEKYLDGIPVMNWLVNQRYVTGSFPRTQDTFVGLKALTKLAEKISPSRNDYTVQLKYK
    KNTKYFNINSEQIDVQNFLEIPEDTKKLEINVGGIGFGLLEVIYQFDLNLVNFEHRFKLD
    LEKQNTGSDYELRLRVCANYIPELTDSQSNMALIEVTLPSGYVVDRNPISEQTTVNPIQN
    MEIRYGGTSVVLYYYKMGTERNCFTVTAYRRFKVALKRPAYVVVYDYYNTNLNAIKVYEV
    DKQNVCEICEEEDCPAECKK""".split()
            ),
        ],
        "SMILES of substrate": ["Negative" for _ in range(5)],
        "Substrate (including stereochemistry)": ["Negative" for _ in range(5)],
        "Type (mono, sesq, di, …)": ["Negative" for _ in range(5)],
    }
    negatives2 = pd.DataFrame(
        {
            col: col_2_vals.get(col, ["Unknown" for _ in range(5)])
            for col in df_main.columns
        }
    )
    df_main = pd.concat((df_main, negatives2))

    # canonical SMILES
    df_main["SMILES_substrate_canonical_no_stereo"] = df_main[
        "SMILES of substrate"
    ].map(get_canonical_smiles)

    df_main["SMILES_product_canonical_no_stereo"] = "Unknown"
    bool_idx = df_main["SMILES of product (including stereochemistry)"].map(
        lambda x: isinstance(x, str) and "k" not in x and "t" not in x
    )
    df_main.loc[bool_idx, "SMILES_product_canonical_no_stereo"] = df_main.loc[
        bool_idx, "SMILES of product (including stereochemistry)"
    ].map(get_canonical_smiles)

    # fixing multi-molecule substrates
    df_main.loc[
        (df_main["Type (mono, sesq, di, …)"] == "tetra")
        & (
            df_main["SMILES_substrate_canonical_no_stereo"]
            == "CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"
        ),
        "SMILES_substrate_canonical_no_stereo",
    ] = "CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O.CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"

    df_main.loc[
        (df_main["Type (mono, sesq, di, …)"] == "tetra")
        & (
            df_main["SMILES_substrate_canonical_no_stereo"]
            == "CC(C)CCCC(C)CCCC(C)CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"
        ),
        "SMILES_substrate_canonical_no_stereo",
    ] = "CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O.CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"

    df_main.loc[
        (df_main["Type (mono, sesq, di, …)"] == "tetra-int")
        & (
            df_main["SMILES_substrate_canonical_no_stereo"]
            == "CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"
        ),
        "SMILES_substrate_canonical_no_stereo",
    ] = "CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O.CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"

    df_main.loc[
        (df_main["Type (mono, sesq, di, …)"] == "tri")
        & (
            df_main["SMILES_substrate_canonical_no_stereo"]
            == "CC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"
        ),
        "SMILES_substrate_canonical_no_stereo",
    ] = "CC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O.CC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"

    df_main.loc[
        (df_main["Type (mono, sesq, di, …)"] == "tri-int")
        & (
            df_main["SMILES_substrate_canonical_no_stereo"]
            == "CC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"
        ),
        "SMILES_substrate_canonical_no_stereo",
    ] = "CC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O.CC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"

    # checking the number of carbons in substrate and product

    def get_carbon_number(smiles: str) -> int:
        """
        Function to calculate the number of carbon atoms in a molecule represented by a SMILES string.

        :param smiles: The SMILES string representing the molecule.

        :return: The number of carbon atoms in the molecule.
        """

        m = Chem.MolFromSmiles(smiles)
        patt = Chem.MolFromSmarts("[C]")
        pm = m.GetSubstructMatches(patt)
        return len(pm)

    # original Jung's functions
    def check_if_c_equal_in_2lists(string_list1: str, string_list2: str) -> np.ndarray:
        """
        Function to check if the number of carbon atoms (represented by the character 'c')
        is equal in two lists of strings.

        :param string_list1: The first list of strings.
        :param string_list2: The second list of strings.

        :return: A numpy array of boolean values indicating whether the count of 'c' is equal
                 in the corresponding strings from both lists.
        """

        ok_flag_list = []

        for s1, s2 in zip(string_list1, string_list2):
            if isinstance(s1, float) or isinstance(s2, float):
                ok_flag_list.append(False)
            else:
                c1, c2 = s1.lower().count("c"), s2.lower().count("c")
                ok_flag_list.append(c1 == c2)
        return np.array(ok_flag_list)

    valid_equal_c_np = check_if_c_equal_in_2lists(
        df_main["SMILES_product_canonical_no_stereo"].values,
        df_main["SMILES_substrate_canonical_no_stereo"].values,
    )
    df_main = df_main.loc[valid_equal_c_np]

    # cleaning the sequences
    df_main = df_main.loc[
        df_main["Amino acid sequence"].map(lambda x: not isinstance(x, float))
    ]
    df_main["Amino acid sequence"] = df_main["Amino acid sequence"].map(
        lambda x: x.replace("*", "").replace('"', "").replace("'", "")
    )
    df_main["Amino acid sequence"] = df_main["Amino acid sequence"].map(
        lambda x: "".join(x.split())
    )

    # removing rows with crucial values missing, preserving full sequences only

    df_main["OK row flag"] = (
        (
            ~df_main[
                [
                    "Amino acid sequence",
                    "SMILES_substrate_canonical_no_stereo",
                ]
            ].isnull()
        )
        .any(axis=1)
        .astype(int)
    )
    df_main.loc[
        df_main["Fragment"].map(lambda x: isinstance(x, str) and x.lower() == "yes"),
        "OK row flag",
    ] = 0
    df_main = df_main[df_main["OK row flag"] == 1]

    df_main["is_substrate_predicted"] = 1

    # deriving clean kingdom info
    df_main["Kingdom"] = df_main["Kingdom (plant, fungi, bacteria)"]
    for animal_init_category in [
        "Animalia",
        "Coral",
        "Insecta",
        "Animalia_(Marine_Sponge)",
        "Human",
        "Human_(Animalia)",
    ]:
        df_main.loc[df_main["Kingdom"] == animal_init_category, "Kingdom"] = "Animals"

    df_main.loc[df_main["Kingdom"] == "Slime mold", "Kingdom"] = "Protists"
    df_main.loc[df_main["Kingdom"] == "Red_algae", "Kingdom"] = "Plants"
    df_main.loc[df_main["Kingdom"] == "Plantae", "Kingdom"] = "Plants"
    df_main.loc[df_main["Kingdom"] == "Viruses", "Kingdom"] = "Viruses"
    df_main.to_csv("data/TPS-Nov19_2023_verified_all_reactions.csv", index=None)

    # preserving only major reactions
    df_products_count = (
        df_main.groupby(["Uniprot ID", "SMILES_substrate_canonical_no_stereo"])[
            "SMILES_product_canonical_no_stereo"
        ]
        .count()
        .reset_index()
    )

    single_product_ids = set(
        df_products_count.loc[
            df_products_count["SMILES_product_canonical_no_stereo"] == 1,
            ["Uniprot ID", "SMILES_substrate_canonical_no_stereo"],
        ]
        .apply(
            lambda x: f'{x["Uniprot ID"]}_{x["SMILES_substrate_canonical_no_stereo"]}',
            axis=1,
        )
        .values
    )

    df_main["Multi-product TPS reaction"] = 1
    df_main.loc[
        df_main[["Uniprot ID", "SMILES_substrate_canonical_no_stereo"]]
        .apply(
            lambda x: f'{x["Uniprot ID"]}_{x["SMILES_substrate_canonical_no_stereo"]}',
            axis=1,
        )
        .isin(single_product_ids),
        "Multi-product TPS reaction",
    ] = 0

    major_product_data_bool_idx = (
        (df_main["Multi-product TPS reaction"] == 0)
        & (df_main["Product is major"].str.lower() != "no")
    ) | (
        (df_main["Multi-product TPS reaction"] == 1)
        & (df_main["Product is major"].str.lower() == "yes")
    )

    df_main.drop("Multi-product TPS reaction", axis=1, inplace=True)

    df_main["product_is_major"] = 0
    df_main.loc[major_product_data_bool_idx, "product_is_major"] = 1

    id_2_substrate_counts = (
        df_main.loc[~major_product_data_bool_idx]
        .groupby("Uniprot ID")["SMILES_substrate_canonical_no_stereo"]
        .nunique()
    )

    id_with_single_substrate = set(
        id_2_substrate_counts[id_2_substrate_counts == 1].index
    )

    df_main["is_substrate_predicted"] = 0
    df_main.loc[
        major_product_data_bool_idx
        | df_main["Uniprot ID"].map(lambda x: x in id_with_single_substrate),
        "is_substrate_predicted",
    ] = 1

    substrates_series = df_main.groupby("Uniprot ID")[
        "SMILES_substrate_canonical_no_stereo"
    ].agg(lambda x: tuple(sorted(set(x))))

    substrates_combo_counts = substrates_series.value_counts()
    supported_substrates = set()
    for substr_combo in substrates_combo_counts.index[substrates_combo_counts >= 10]:
        for substr in substr_combo:
            supported_substrates.add(substr)

    df_main = df_main[df_main["is_substrate_predicted"] == 1]

    # storing the final dataset
    df_main.to_csv("data/TPS-Nov19_2023_verified.csv", index=None)

    # preparing experimental validation data
    df_wetlab_hard = pd.read_csv("data/seqs_to_predict_wet_hard.csv").dropna()
    df_wetlab_easy = pd.read_csv("data/seqs_to_predict_wet_easy.csv").dropna()

    df_wetlab_easy["CC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"] = [
        1,
        1,
        0,
        0,
        0,
        1,
    ]
    df_wetlab_easy["CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"] = [
        0,
        0,
        1,
        1,
        1,
        0,
    ]
    df_wetlab_easy["Kingdom"] = [
        "Fungi",
        "Bacteria",
        "Plants",
        "Plants",
        "Plants",
        "Plants",
    ]
    df_wetlab_easy["protein_signatures"] = "With"

    df_wetlab_hard["CC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"] = [
        1,
        0,
        0,
        1,
        1,
        1,
        1,
    ]
    df_wetlab_hard["CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O"] = [
        0,
        1,
        1,
        0,
        0,
        0,
        1,
    ]
    df_wetlab_hard["Kingdom"] = [
        "Viruses",
        "Archaea",
        "Bacteria",
        "Bacteria",
        "Bacteria",
        "Archaea",
        "Archaea",
    ]
    df_wetlab_hard["protein_signatures"] = "Without"
    df_wetlab = pd.concat((df_wetlab_easy, df_wetlab_hard))
    df_wetlab_long = pd.melt(
        df_wetlab,
        id_vars=["ID", "Amino acid sequence", "Kingdom", "protein_signatures"],
        var_name="SMILES_substrate_canonical_no_stereo",
        value_vars=[
            "CC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O",
            "CC(C)=CCCC(C)=CCCC(C)=CCCC(C)=CCOP([O-])(=O)OP([O-])([O-])=O",
        ],
    )
    df_wetlab_long = df_wetlab_long[df_wetlab_long["value"] == 1]
    df_wetlab_long.drop("value", axis=1, inplace=True)
    # storing the wetlab dataset
    df_wetlab_long.to_csv("data/df_wetlab_long_clean.csv", index=None)

    # computing categories per kingdom
    id_2_kingdom_dataset = (
        df_main[["Uniprot ID", "Kingdom"]].set_index("Uniprot ID").to_dict()["Kingdom"]
    )
    with open("data/id_2_kingdom_dataset.pkl", "wb") as file:
        pickle.dump(id_2_kingdom_dataset, file)

    # computing categories per kingdom for the wetlab dataset
    id_2_kingdom_wetlab_dataset = (
        df_wetlab_long[["ID", "Kingdom"]].set_index("ID").to_dict()["Kingdom"]
    )
    with open("data/id_2_kingdom_wetlab_dataset.pkl", "wb") as file:
        pickle.dump(id_2_kingdom_wetlab_dataset, file)

    # computing categories per protein signature presence
    uniprot_families_df = pd.read_csv("data/uniprot_info_dataset.tsv", sep="\t")
    uniprot_families_df["domains"] = "With"
    uniprot_families_df.loc[
        uniprot_families_df.apply(
            lambda row: np.all(
                list(
                    map(
                        lambda x: isinstance(x, float),
                        row[["Pfam", "SUPFAM", "InterPro"]].values,
                    )
                )
            ),
            axis=1,
        ),
        "domains",
    ] = "Without"
    id_2_domains_presence = (
        uniprot_families_df[["From", "domains"]].set_index("From").to_dict()["domains"]
    )
    with open("data/id_2_domains_presence.pkl", "wb") as file:
        pickle.dump(id_2_domains_presence, file)

    # computing categories per protein signature presence for the wetlab dataset
    id_2_domains_presence_wetlab_dataset = (
        df_wetlab_long[["ID", "protein_signatures"]]
        .set_index("ID")
        .to_dict()["protein_signatures"]
    )
    with open("data/id_2_domains_presence_wetlab_dataset.pkl", "wb") as file:
        pickle.dump(id_2_domains_presence_wetlab_dataset, file)
